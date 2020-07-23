import requests 
import json
from openpyxl import load_workbook
import argparse

ap = argparse.ArgumentParser()
ap.add_argument('-q', '--query', required = True,
help= "The excel file name to run this script on(please dont forget to add the extension name)")
args = vars(ap.parse_args())

with open('Weather-key.txt', 'r') as f:
    key = f.read()

while True:

    b = load_workbook(filename=args['query'], read_only=False)
    sheet = b.active

    def fetch_temp(city, apikey, unit):
        params = {
            'q': city,
            'format': 'json',
            'key': apikey,
            'fx': 'no'
            }
        base_url= 'http://api.worldweatheronline.com/premium/v1/weather.ashx'    
        response = requests.get(base_url, params = params)
        data = response.json()
        if unit == 'C':
            return data['data']['current_condition'][0]['temp_C']
        if unit == 'F':
            return data['data']['current_condition'][0]['temp_F']

    #Unit and update or not will be gathered from the excel sheet
    temps = []
    updates = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        temps.append(fetch_temp(row[0],key,row[2]))
        updates.append(row[3])
    for i in range(1, len(temps)+1):
        if updates[i-1] == 1 :
            sheet["B"+ str(i+1)] = int(temps[i-1])
        else:
            continue
    b.save(filename=args['query'])
 
        
